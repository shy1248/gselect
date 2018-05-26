#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
@Author: yushuibo
@Copyright (c) 2018 yushuibo. All rights reserved.
@Licence: GPL-2
@Email: hengchen2005@gmail.com
@File: tmp.py
@Create: 2018-05-26 07:30:47
@Last Modified: 2018-05-26 07:30:47
@Desc: --
"""

import openpyxl

from server import *


def rs2str(rs, col):
    rows = []
    for row in rs:
        dest_col = str(row[col])
        if dest_col:
            rows.append(str(row[col]))
    return '(' + ', '.join(rows) + ')'

def rs2list(rs):
    rows = []
    for row in rs:
        rows.append(list(row))
    return rows

def openxl():
    return openpyxl.Workbook()

def savexl(wb, path):
    wb.save(path)

def wte2xl(wb, titles, data, sheet, sheet_indx):
    ws = wb.create_sheet(sheet, sheet_indx)

    for i in range(0, len(titles)):
        ws.cell(row=1, column=i+1, value=titles[i])

    for r in range(1, len(data)+1):
        for c in range(0, len(data[r-1])):
            ws.cell(row=r+1, column=c+1, value=str(data[r-1][c]))




login_playerid = ZgmjLog().select('SELECT DISTINCT playerid FROM chess_login_2018_05_25')


bind_code = ZgmjGame().select('SELECT id, bingAgentId FROM `tb_player` WHERE id IN {player_id}'.format(player_id=rs2str(login_playerid, 0)))

agent_info = ZgmjAgent().select('SELECT a.login_id, a.`level`, a.`name`, b.login_id, b.`name`, a.bind_code FROM `agent` a LEFT JOIN agent b ON a.parent_id=b.id WHERE a.bind_code IN {bind_id}'.format(bind_id=rs2str(bind_code, 1)))

wb = openxl()
wte2xl(wb, ['玩家ID', '绑定ID'], rs2list(bind_code), 'sheet1', 0)
wte2xl(wb, ['绑定代理', '绑定代理级别', '绑定代理名称', '父级代理', '父级代理名称','绑定ID'], rs2list(agent_info), 'sheet2', 1)
savexl(wb, '/Users/yu/Desktop/20180525.xlsx')



