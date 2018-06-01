#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
@Author: yushuibo
@Copyright (c) 2018 yushuibo. All rights reserved.
@Licence: GPL-2
@Email: hengchen2005@gmail.com
@Create: gselect.py
@Last Modified: 2018/5/25 下午 05:28
@Desc: --
"""

import copy
import os
import re
from optparse import OptionParser

import openpyxl
import yaml

from server import *

template_pattern = re.compile(r'.*%\((.*)\)s.*')


def is_subset(l1, l2):
    """judge list2 is a subset of list1 or not"""
    if not l2:
        return True, None

    tmp = copy.deepcopy(l1)
    for i in l2:
        try:
            tmp.remove(i)
        except ValueError:
            return False, i
    del tmp

    return True, None


def autorename(path):
    name, ext = os.path.splitext(path)
    return '{}_{}{}'.format(name, str(round(time.time() * 1000)), ext)


def args_parse():
    usage = 'Usage: %prog -p PLAYBOOK'
    parser = OptionParser(usage, version='%prog 1.0')
    parser.add_option(
        '-p',
        '--playbook',
        dest='playbook',
        help="Parse statment from a playbook")

    (options, args) = parser.parse_args()

    return options.playbook
    # if len(args) != 2:
    #     parser.error('Incorrect number of arguments!')


class Statment(object):
    _ids = {}

    def __init__(self, ser_, id__, based_, names_, sql_, save_):
        self.ser = ser_
        self.id_ = id__
        self.based = based_
        self.names = names_
        self.sql = sql_
        self.save = save_
        self.result = None

        if self.id_ not in Statment._ids:
            Statment._ids[self.id_] = self
        else:
            raise RuntimeError('Duplicated statment id \'{}\''.format(
                self.id_))

        if self.id_ == self.based:
            raise RuntimeError(
                'The statment \'{}\' can not based itself!'.format(self.id_))

    def gettitles(self):
        return list(self.names.keys())

    def getalias(self):
        return list(self.names.values())

    def has_template(self):
        """judge a sql contant a template or not"""
        return True if template_pattern.match(self.sql) else False

    def template_keys(self):
        """get template keys from a sql statment"""
        return template_pattern.findall(self.sql)

    def template_values(self):
        """get template valus from result list"""

        tkeys = self.template_keys()

        for _stmt in stmts:
            # if can be found result by bsaed id
            if self.based == _stmt.id_:

                if not _stmt.result:
                    raise RuntimeError(
                        'The statment id \'{}\' must be excuted before when '
                        'another statment based it.'.format(self.based))

                data = _stmt.result
                titles = _stmt.gettitles()
                b, i = is_subset(titles, tkeys)
                # template key cound not found in the result
                if not b:
                    raise RuntimeError(
                        'Template key \'{}\' is not found at the resultset.'
                        .format(i))

                for key in tkeys:
                    ind = titles.index(key)
                    values = []

                    for m in range(0, len(data)):
                        value = data[m][ind]
                        # filte the column who's value is null or ''
                        if value:
                            values.append(str(value))
                    values_str = ', '.join(values)
                    del values
                    yield {key: values_str}

                break

        # if not found based statment
        else:
            raise RuntimeError(
                'The statment id \'{}\' based statment not found.'.format(
                    self.id_))

    def fmtsql(self):
        tvalues = self.template_values()
        for value in tvalues:
            self.sql = self.sql % value

    def exec_(self):
        print('==> Starting execute statment --> {} ...'.format(
            self.sql[:100]))
        # get a class order by a string
        server = globals().get(self.ser)()
        if not server:
            raise RuntimeError('Server \'{}\' is not defined.'.format(
                self.ser))

        rs = server.do_select(self.sql)
        self.result = rs


class Handler(object):
    def __init__(self, path):
        self.path = path
        self.wb = openpyxl.Workbook()
        self.st_ind = 0

    def __save(self):
        if os.path.exists(self.path):
            new = autorename(self.path)
            print('File \'{}\' already exist, use new file \'{}\''.format(
                self.path, new))
            self.path = new

        self.wb.save(self.path)

    def __w_stmt(self, _stmt):
        # add titles
        _stmt.result.insert(0, _stmt.getalias())
        data = _stmt.result
        ws = self.wb.create_sheet()
        for r in range(0, len(data)):
            for c in range(0, len(data[r])):
                ws.cell(row=r + 1, column=c + 1, value=str(data[r][c]))

    def write(self, _stmts):
        print('==> Start write result to file \'{}\' ...'.format(self.path))
        # delete the default sheet
        self.wb.remove(self.wb.active)
        for _stmt in _stmts:
            if _stmt.save:
                self.__w_stmt(_stmt)

        self.__save()

        print('==> Save done, enjoy!')


if __name__ == '__main__':
    file = args_parse()
    with open(file, 'r', encoding='UTF-8') as f:
        conf = yaml.load(f)
    # pase playbook
    stmts = []
    print('==> Parser config file ...')
    for select in conf['selects']:
        ser = select['server']

        for statment in select['statments']:
            id_ = statment['id']
            based = statment['based']
            names = statment['names']
            sql = statment['sql']
            save = statment['save']

            print('==> Found select statment:')
            print('\t-> server: {}'.format(ser))
            print('\t-> id: {}'.format(id_))
            print('\t-> based: {}'.format(based))
            print('\t-> names: {}'.format(names))
            print('\t-> sql: {}'.format(sql))

            stmt = Statment(ser, id_, based, names, sql, save)
            stmts.append(stmt)
    # execute statment
    for stmt in stmts:
        if stmt.id_ == 0 and stmt.based is not None:
            raise RuntimeError(
                'The fisrt sql statment cannot be based another statment!')
        elif stmt.id_ == 0 and stmt.has_template():
            raise RuntimeError(
                'The SQL in the fisrt statment cannot used templates!')
        elif stmt.id_ != 0 and stmt.has_template() and stmt.based is None:
            raise RuntimeError('Sql statment \'{}\', based id is None.'.format(
                stmt.sql))
        elif stmt.id_ != 0 and stmt.has_template():
            sql = stmt.fmtsql()
            stmt.exec_()
        else:
            stmt.exec_()

    print('==> All select statments finshed.')
    Handler(conf['handler']).write(stmts)
